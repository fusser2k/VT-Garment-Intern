"""
Cut Planning Model v2 — Page 1: Data Extraction
=================================================
This is a NEW, separate model from the original cut_plan_model project.
Page 1's job is to correctly read an input Excel file and extract it into
the 17-field schema below, regardless of what the source file's actual
column headers happen to be.

Later pages (e.g. an actual cut-planning/table-selection page) can build on
top of the DataFrame this module produces.
"""

from datetime import datetime, timedelta, date
from typing import Optional, List, Tuple
from zoneinfo import ZoneInfo

import json
import math
import os
import re
import unicodedata
from collections import Counter

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"

# This tool is for a Thai garment manufacturer - every timestamp shown to
# the user (extraction time, generation time, filenames' dates, etc.) should
# reflect Thailand's local time, regardless of what timezone the SERVER
# itself happens to be running in (most cloud hosts, including Railway,
# default their containers to UTC). Falls back to a fixed UTC+7 offset if
# the IANA timezone database isn't available in this environment (e.g. a
# minimal container missing tzdata) - see requirements.txt for the tzdata
# package that normally prevents this fallback from ever being needed.
try:
    TH_TZ = ZoneInfo("Asia/Bangkok")
except Exception:
    from datetime import timezone
    TH_TZ = timezone(timedelta(hours=7))


def now_th() -> datetime:
    """Current date/time in Thailand (Asia/Bangkok, UTC+7), regardless of
    the server's own system timezone. Use this everywhere a timestamp is
    shown to the user or used to compute "today" - never use the bare
    datetime.now() (server-local time) for anything user-facing."""
    return datetime.now(TH_TZ)


# The 17 fields Page 1 must extract/compute, in display order.
OUTPUT_COLUMNS = [
    "Sewing Line",
    "JobCut - Suffix",
    "Table ID",
    "Colorway",
    "Mark Type",
    "Layer",
    "Qty",
    "Qty Complete",
    "Difference",
    "% Complete",
    "Status",
    "Sewing Target Per Day (with OT)",
    "Sewing Target Per Day (no OT)",
    "Table No. (Mark Type 101)",
    "Decoration",
    "FG Start Date",
    "Start Cut",
]

# Columns in OUTPUT_COLUMNS that are always COMPUTED from another extracted
# column rather than read/aliased from the source file - never counted as
# "missing" on their own, since they don't come from the file at all.
COMPUTED_COLUMNS = ["Sewing Target Per Day (no OT)"]

# Direct aliases: source header -> one of the OUTPUT_COLUMNS above.
# Add more entries here as real production files use different header text.
DIRECT_ALIASES = {
    "Sewing Line": "Sewing Line",
    "No. of sewing line": "Sewing Line",
    "JobCut - Suffix": "JobCut - Suffix",
    "Table ID": "Table ID",
    "Table No.": "Table ID",
    "Colorway": "Colorway",
    "Mark Type": "Mark Type",
    "Layer": "Layer",
    "Qty": "Qty",
    "Total": "Qty",
    "Qty Complete": "Qty Complete",
    "Difference": "Difference",
    "% Complete": "% Complete",
    "Status": "Status",
    "Sewing Target Per Day (with OT)": "Sewing Target Per Day (with OT)",
    "Sewing Target Per Day": "Sewing Target Per Day (with OT)",
    "Sewing target per day": "Sewing Target Per Day (with OT)",
    "Table No. (Mark Type 101)": "Table No. (Mark Type 101)",
    "Decoration": "Decoration",
    "FG Start Date": "FG Start Date",
    "Start Cut": "Start Cut",
}

# Columns that, together, can build "JobCut - Suffix" when the source file
# keeps Job Cut and Suffix as two separate columns instead of one combined one.
JOBCUT_COL_CANDIDATES = ["Job Cut", "JobCut", "Job Order"]
SUFFIX_COL_CANDIDATES = ["Suffix"]


def load_input(path_or_buffer) -> Tuple[pd.DataFrame, List[str], int]:
    """Read an input Excel file and extract it into the OUTPUT_COLUMNS schema.

    Unlike a strict schema check, this does NOT raise an error if some of
    the 17 fields aren't present in the source file — those fields are
    simply returned blank, and their names are included in the returned
    `missing_columns` list so the caller (web page / CLI) can tell the user
    which fields the source file didn't provide.

    Rows whose Sewing Line does not start with "VS" are filtered out (the
    planning team only tracks VS-prefixed sewing lines, e.g. VSEW012,
    VS02+06). The number of rows dropped this way is returned as
    `filtered_out_count`.

    Returns (extracted_df, missing_columns, filtered_out_count).
    """
    raw = pd.read_excel(path_or_buffer)
    raw_cols = set(raw.columns)
    n = len(raw)

    extracted = {}
    missing_columns = []

    # --- JobCut - Suffix: combined column, or built from two separate ones ---
    if "JobCut - Suffix" in raw_cols:
        extracted["JobCut - Suffix"] = raw["JobCut - Suffix"].astype(str)
    else:
        jobcut_col = next((c for c in JOBCUT_COL_CANDIDATES if c in raw_cols), None)
        suffix_col = next((c for c in SUFFIX_COL_CANDIDATES if c in raw_cols), None)
        if jobcut_col is not None:
            if suffix_col is not None:
                extracted["JobCut - Suffix"] = (
                    raw[jobcut_col].astype(str) + "-" + raw[suffix_col].astype(str)
                )
            else:
                extracted["JobCut - Suffix"] = raw[jobcut_col].astype(str)
        else:
            extracted["JobCut - Suffix"] = pd.Series([""] * n)
            missing_columns.append("JobCut - Suffix")

    # --- Everything else: direct alias lookup (skip computed columns -
    # those are filled in afterward, not read from the source file at all) ---
    for target_col in OUTPUT_COLUMNS:
        if target_col == "JobCut - Suffix" or target_col in COMPUTED_COLUMNS:
            continue

        source_col = next(
            (src for src, tgt in DIRECT_ALIASES.items() if tgt == target_col and src in raw_cols),
            None,
        )
        if source_col is not None:
            extracted[target_col] = raw[source_col]
        else:
            extracted[target_col] = pd.Series([None] * n)
            missing_columns.append(target_col)

    # --- Computed columns ---
    # Sewing Target Per Day (no OT) = Sewing Target Per Day (with OT) scaled
    # down from a 10.75-hour (with overtime) workday to a 7.75-hour (regular,
    # no overtime) workday. Not counted as "missing" even when the source
    # file lacks a column for it, since it's never meant to come from the
    # file - only ends up blank if the WITH-OT value it's computed from is
    # itself missing/non-numeric.
    with_ot_numeric = pd.to_numeric(extracted["Sewing Target Per Day (with OT)"], errors="coerce")
    no_ot = with_ot_numeric / 10.75 * 7.75
    extracted["Sewing Target Per Day (no OT)"] = no_ot.round().astype("Int64")

    df = pd.DataFrame(extracted, columns=OUTPUT_COLUMNS)

    # --- Filter: only keep rows whose Sewing Line starts with "VS" ---
    before_count = len(df)
    sewing_line_str = df["Sewing Line"].astype(str).str.strip()
    df = df[sewing_line_str.str.startswith("VS")].reset_index(drop=True)
    filtered_out_count = before_count - len(df)

    return df, missing_columns, filtered_out_count


TITLE_FONT_COLOR = "1F4E78"
WARN_FONT_COLOR = "C00000"


def write_extracted_workbook(
    df: pd.DataFrame,
    missing_columns: List[str],
    output_path: str,
    generated_at: Optional[datetime] = None,
    filtered_out_count: int = 0,
) -> None:
    """Write the formatted Page 1 output workbook: just the Extracted Data
    sheet (single-sheet workbook), starting directly with the header row -
    no title/notice row above it. missing_columns/filtered_out_count/
    generated_at are still accepted (existing callers pass them) but no
    longer written into the sheet itself; that information is shown in the
    web UI's own banner instead."""
    generated_at = generated_at or now_th()

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
    cell_font = Font(name=FONT_NAME, size=10)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # ---- Sheet 1: Extracted Data ----
    ws = wb.active
    ws.title = "Extracted Data"

    header_row = 1

    headers = list(df.columns)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val if pd.notna(val) else None)
            c.font = cell_font
            c.alignment = center
            c.border = border

    widths = [12, 16, 10, 12, 10, 8, 8, 12, 11, 11, 10, 20, 20, 20, 12, 14, 14]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = f"A{header_row + 1}"
    last_row = max(len(df) + header_row, header_row)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"

    wb.save(output_path)


def run(input_path: str, output_path: str):
    """Convenience entry point: extract, write, return (df, missing_columns, filtered_out_count)."""
    df, missing_columns, filtered_out_count = load_input(input_path)
    write_extracted_workbook(df, missing_columns, output_path, filtered_out_count=filtered_out_count)
    return df, missing_columns, filtered_out_count


WIP_HEADER_MARKER = "ไลน์เย็บ"
# Some WIP report files use the Thai header text above; others (seen in
# later template revisions) use the English "Sewing line" instead - accept
# either, matched case-insensitively.
WIP_HEADER_MARKER_VARIANTS = {WIP_HEADER_MARKER.strip().lower(), "sewing line"}
WIP_STOP_MARKER = "ไลน์"  # marks the start of the unrelated table-status mini-table further down the sheet

# Maps this WIP template's Excel column letters to friendly English field
# names. The sheet has TWO header blocks (one per set of sewing lines) that
# repeat the exact same column layout, so one static mapping covers both.
WIP_COLUMN_MAP = {
    "A": "Sewing Line Code",
    "B": "Sewing Line",
    "C": "Target Hours (OTP 100 Plan)",
    "D": "OTP Sewing Ratio",
    "E": "Target per Hour",
    "F": "Target for Day (with OT)",
    "G": "Target for Day (without OT)",
    "H": "Target Morning",
    "I": "Target Afternoon",
    "J": "Target OT",
    "K": "Actual Morning",
    "L": "Actual Morning (Extra)",
    "M": "Actual Afternoon",
    "N": "Actual Afternoon (Extra)",
    "O": "Actual OT",
    "Q": "Morning Target Shortfall",
    "R": "WIP in Sewing Line",
    "S": "Sewn Yesterday",
    "T": "Bundled Waiting for Sewing (Full Day)",
    "U": "WIP Cut Waiting to be Bundled (4-12 hrs)",
    "V": "Waiting to Cut",
    "W": "Lead Time Morning (hrs)",
    "X": "Lead Time Morning Status",
    "Y": "Lead Time Afternoon (hrs)",
    "Z": "Lead Time Afternoon Status",
    "AA": "Lead Time OT (hrs)",
    "AB": "Lead Time OT Status",
    "AC": "Reason Missed Morning Target",
    "AD": "Detail Morning",
    "AE": "Afternoon Target Shortfall",
    "AF": "Reason Missed Afternoon Target",
    "AG": "Detail Afternoon",
    "AH": "OT Target Shortfall (Full Day)",
    "AI": "Reason Missed OT Target",
    "AJ": "Detail OT",
    "AK": "Cause Category",
    "AL": "Cut Waiting to be Bundled",
    "AM": "Add: Cut Fully Bundled",
    "AN": "Add: WIP Bundled Waiting for Sewing",
    "AO": "OTP Sewing (Reference Date)",
    "AP": "OT Total",
    "AQ": "Total without OT",
}

WIP_COLUMNS = list(WIP_COLUMN_MAP.values())

# Column U's conditional-formatting thresholds in the source file (Excel's
# own red/green/orange coloring on that cell) - used to derive a text field
# since, unlike the Lead Time columns, this one has no adjacent status
# column already spelling out what the color means.
WIP_CUT_WAITING_LOW = 3.5
WIP_CUT_WAITING_HIGH = 12.5


def classify_wip_cut_waiting(value) -> str:
    """Mirror the source file's conditional formatting on column U ("WIP
    ตัดรอจัดงาน 4-12 ชม"): red = underproduction (too little cut fabric
    queued up waiting to be bundled), green = neutral/healthy, orange =
    overproduction (too much queued up)."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return ""
    if v < WIP_CUT_WAITING_LOW:
        return "Underproduction"
    if v > WIP_CUT_WAITING_HIGH:
        return "Overproduction"
    return "Neutral"


# Insert the derived health field right after the raw value it's based on.
_insert_at = WIP_COLUMNS.index("WIP Cut Waiting to be Bundled (4-12 hrs)") + 1
WIP_COLUMNS.insert(_insert_at, "WIP Cut Waiting Health")

# Lead Time (W/Y/AA) conditional-formatting thresholds. The source file uses
# TWO different threshold sets depending on which block of sewing lines a
# row belongs to: individually-named lines (the first block, tighter
# 9.5-16.5 "healthy" band) vs. merged lines like "VS02+06" (the second
# block, wider 15.5-32.5 band, since a merged line naturally carries more
# backlog). load_wip() tracks which block each row came from and picks the
# matching thresholds automatically.
LEAD_TIME_INDIVIDUAL_LOW = 9.5
LEAD_TIME_INDIVIDUAL_HIGH = 16.5
LEAD_TIME_MERGED_LOW = 15.5
LEAD_TIME_MERGED_HIGH = 32.5


def classify_lead_time(value, is_merged_block: bool = False) -> str:
    """Mirror the source file's conditional formatting on the Lead Time
    Morning/Afternoon columns (W/Y): red = underproduction (line is running
    low on queued work - at risk of sitting idle), green =
    neutral/healthy, orange = overproduction (too much backlog queued up)."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return ""
    low, high = (LEAD_TIME_MERGED_LOW, LEAD_TIME_MERGED_HIGH) if is_merged_block else (
        LEAD_TIME_INDIVIDUAL_LOW, LEAD_TIME_INDIVIDUAL_HIGH
    )
    if v < low:
        return "Underproduction"
    if v > high:
        return "Overproduction"
    return "Neutral"


for _col in ["Lead Time Morning (hrs)", "Lead Time Afternoon (hrs)"]:
    _pos = WIP_COLUMNS.index(_col) + 1
    WIP_COLUMNS.insert(_pos, _col.replace("(hrs)", "Health").replace("  ", " ").strip())


def load_wip(path_or_buffer) -> pd.DataFrame:
    """Read a Work-in-Process (WIP) buffer report and extract it into the
    fixed WIP_COLUMNS schema.

    This template repeats its header row once per group of sewing lines
    (e.g. once for the individually-named lines, again for the merged
    lines like "VS02+06"). Rather than assuming a fixed row range, this
    scans column B for the literal header text ("ไลน์เย็บ" or, in newer
    template revisions, the English "Sewing line" - matched
    case-insensitively) to find where
    each block starts, and collects every non-blank row under it as data -
    so it keeps working even if a future day's file has more or fewer
    sewing lines, or an extra block.

    Extraction stops the moment column B hits "ไลน์" (a different, unrelated
    mini-table further down the same sheet that tracks individual cutting
    tables, not sewing lines - the WIP_STOP_MARKER) - so that table is
    correctly left out rather than misread as more sewing-line rows.
    """
    wb = openpyxl.load_workbook(path_or_buffer, data_only=True)
    ws = wb[wb.sheetnames[0]]

    col_indices = {letter: get_column_letter_index(letter) for letter in WIP_COLUMN_MAP}

    rows_out = []
    collecting = False
    block_index = -1  # increments each time a new header block starts
    for r in range(1, ws.max_row + 1):
        b_val = ws.cell(row=r, column=2).value
        b_str = str(b_val).strip() if b_val is not None else ""

        if b_str.lower() in WIP_HEADER_MARKER_VARIANTS:
            collecting = True
            block_index += 1
            continue
        if b_str == WIP_STOP_MARKER:
            break
        if collecting and b_str:
            row_data = {}
            for letter, field_name in WIP_COLUMN_MAP.items():
                row_data[field_name] = ws.cell(row=r, column=col_indices[letter]).value
            row_data["WIP Cut Waiting Health"] = classify_wip_cut_waiting(
                row_data.get("WIP Cut Waiting to be Bundled (4-12 hrs)")
            )
            # First block = individually-named lines (tighter thresholds);
            # any block after that = merged lines like "VS02+06" (wider
            # thresholds), matching this template's actual layout.
            is_merged_block = block_index > 0
            row_data["Lead Time Morning Health"] = classify_lead_time(
                row_data.get("Lead Time Morning (hrs)"), is_merged_block
            )
            row_data["Lead Time Afternoon Health"] = classify_lead_time(
                row_data.get("Lead Time Afternoon (hrs)"), is_merged_block
            )
            # The source file's own "Sewing Line Code" (column A) isn't
            # always filled in - when it's blank or doesn't look like a
            # real code, recognize it instead from this row's own "Sewing
            # Line" Thai name/team text (see match_sewing_line_code()),
            # so the column is still populated either way. A genuine code
            # already present in the file is always kept as-is and never
            # overridden by the Thai-name guess.
            code_raw = row_data.get("Sewing Line Code")
            code_str = str(code_raw).strip() if code_raw is not None else ""
            if not code_str or not code_str.upper().startswith("VS"):
                recognized_code = match_sewing_line_code(row_data.get("Sewing Line"))
                if recognized_code:
                    row_data["Sewing Line Code"] = recognized_code
            rows_out.append(row_data)

    return pd.DataFrame(rows_out, columns=WIP_COLUMNS)


def get_column_letter_index(letter: str) -> int:
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)


def write_wip_workbook(df: pd.DataFrame, output_path: str, generated_at: Optional[datetime] = None) -> None:
    """Write a formatted, downloadable version of the extracted WIP data."""
    generated_at = generated_at or now_th()
    wb = Workbook()
    ws = wb.active
    ws.title = "WIP Data"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
    cell_font = Font(name=FONT_NAME, size=10)
    title_font = Font(name=FONT_NAME, size=10, bold=True, color=TITLE_FONT_COLOR)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value=f"Extracted: {generated_at.strftime('%Y-%m-%d %H:%M')}").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(WIP_COLUMNS))

    header_row = 2
    for j, h in enumerate(WIP_COLUMNS, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val if pd.notna(val) else None)
            c.font = cell_font
            c.alignment = center
            c.border = border

    for j in range(1, len(WIP_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 20

    ws.freeze_panes = f"A{header_row + 1}"
    last_row = max(len(df) + header_row, header_row)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(WIP_COLUMNS))}{last_row}"

    wb.save(output_path)


def load_wip_raw(path_or_buffer) -> pd.DataFrame:
    """Read ANY Excel file with NO assumed schema - kept as a fallback for
    WIP files that don't match the known WIP_COLUMN_MAP template (e.g. a
    different sheet layout). load_wip() above should be tried first."""
    df = pd.read_excel(path_or_buffer)
    return df


# ---------------------------------------------------------------------------
# Tab 3: Cut Plan
# ---------------------------------------------------------------------------


# Sewing Line code -> the "root" Thai name/team label as it actually
# appears at the START of that line's row in Tab 2's WIP report "Sewing
# Line" column (before any trailing "/", "+", extra number, or nothing at
# all). These are two entirely different naming systems for the same
# physical sewing lines - Tab 1 uses codes like "VSEW012", Tab 2 identifies
# lines by the supervisor/team's Thai name instead (e.g. "ราตรี"). There's
# no way to derive one from the other programmatically, so this mapping is
# hardcoded here. Update it if a line's assigned supervisor/team changes.
SEWING_LINE_THAI_ROOTS = {
    "VSEW001": "วารินทร์",
    "VSEW002": "นิภาพร",
    "VSEW003": "พิไลย์",
    "VSEW004": "วงเดือน",
    "VSEW005": "จินตนา",
    "VSEW006": "ชุติมันต์",
    "VSEW007": "ส้ม",
    "VSEW008": "นวลจันทร์",
    "VSEW009": "สัมฤทธิ์",
    "VSEW010": "นิชารัศ",
    "VSEW011": "สุพิศ",
    "VSEW012": "ราตรี",
    "VSEW013": "สายรุ้ง",
    "VSEW014": "รำไพ",
    "VSEW015": "ปัญญาพร",
    "VSEW016": "อรอนงค์",
    "VSEW017": "สำราญ",
    "VSEW018": "VSEW018 ผ้าพันคอ",  # this line's WIP row literally spells out its own code
}

# Merged-line codes (two individual lines sharing one combined sewing
# line): recognized when the row's Sewing Line text contains BOTH
# constituent lines' Thai roots (e.g. "นิภาพร + ชุติมันต์" -> VS02+06,
# since "02" and "06" are VSEW002 and VSEW006's own numbers).
SEWING_LINE_MERGED_ROOTS = {
    "VS02+06": ("VSEW002", "VSEW006"),
    "VS08+11": ("VSEW008", "VSEW011"),
}


def _normalize_thai_text(s) -> str:
    """Collapse whitespace and strip, for tolerant Thai-name matching.
    Also applies Unicode NFC normalization - the same-looking Thai text can
    be encoded with combining marks (tone marks, vowels) stored in a
    different order depending on the source, which otherwise silently
    breaks exact/prefix matching even though the text is visually
    identical."""
    return " ".join(unicodedata.normalize("NFC", str(s)).split())


def match_sewing_line_code(thai_text) -> Optional[str]:
    """Recognize which Tab 1 Sewing Line code (e.g. "VSEW012") a Tab 2 WIP
    row's own "Sewing Line" text (e.g. "ราตรี /", "ราตรี", or "ราตรี 2")
    corresponds to, mainly by its Thai name - tolerant of the name ending
    without a trailing "/", or ending with a number, or nothing extra at
    all after the name.

    Checks merged-line patterns FIRST (most specific - e.g. "นิภาพร +
    ชุติมันต์" must resolve to VS02+06, not be mistaken for just VSEW002
    because its text happens to start with "นิภาพร" too), then falls back
    to a single-line prefix match. Returns None if nothing matches.
    """
    if thai_text is None:
        return None
    normalized = _normalize_thai_text(thai_text)
    if not normalized:
        return None

    for merged_code, (code_a, code_b) in SEWING_LINE_MERGED_ROOTS.items():
        root_a = SEWING_LINE_THAI_ROOTS[code_a]
        root_b = SEWING_LINE_THAI_ROOTS[code_b]
        if root_a in normalized and root_b in normalized:
            return merged_code

    for code, root in SEWING_LINE_THAI_ROOTS.items():
        if normalized.startswith(root):
            return code

    return None


def compute_wip_target_overrides(wip_df: pd.DataFrame) -> dict:
    """From an already-extracted Tab 2 WIP DataFrame (see load_wip()),
    build {Sewing Line code: Target for Day (with OT)}.

    Reads the "Sewing Line Code" column directly - load_wip() already
    backfills that column via Thai-name recognition (see
    match_sewing_line_code() / SEWING_LINE_THAI_ROOTS above) for any row
    where the source file's own column A was blank or unusable, so by the
    time this function runs every row that CAN be identified already has a
    code here, whether it came from the file itself or was recognized from
    its Thai name.

    Lines with no usable code, or no numeric target value, are simply
    absent from the returned dict - callers should fall back to Tab 1's
    own Sewing Target Per Day (with OT) for those.

    Returns {} if wip_df is None, empty, or missing the expected columns
    (e.g. an unstructured/raw-preview WIP file) - never raises.
    """
    if wip_df is None or len(wip_df) == 0:
        return {}
    if "Sewing Line Code" not in wip_df.columns or "Target for Day (with OT)" not in wip_df.columns:
        return {}

    overrides = {}
    for _, row in wip_df.iterrows():
        code_raw = row.get("Sewing Line Code")
        if code_raw is None:
            continue
        code = str(code_raw).strip()
        if not code or not code.upper().startswith("VS"):
            continue
        val = pd.to_numeric(row["Target for Day (with OT)"], errors="coerce")
        if pd.notna(val):
            overrides[code] = float(val)

    return overrides


def compute_wip_session_targets(wip_df: pd.DataFrame) -> dict:
    """From an already-extracted Tab 2 WIP DataFrame (see load_wip()),
    build {Sewing Line code: (Target Morning, Target Afternoon, Target OT)}
    - the per-session targets Tab 3's planning is built around (see
    build_cut_plan()). Matched the same way as compute_wip_target_overrides()
    (via the "Sewing Line Code" column, which load_wip() already backfills
    via Thai-name recognition when the source file's own column A is
    blank).

    A line WITH a matching row in the WIP data, but a missing/non-numeric
    value for just one of the three target columns, gets 0 for that
    specific missing piece while keeping the other two real values (never
    raises). A line with NO matching row at all simply has no entry in the
    returned dict - build_cut_plan() treats that as "no Tab 2 coverage" and
    skips planning that Sewing Line entirely, rather than guessing with a
    (0, 0, 0) fallback.

    Returns {} if wip_df is None, empty, or missing the expected columns
    (e.g. an unstructured/raw-preview WIP file) - never raises.
    """
    if wip_df is None or len(wip_df) == 0:
        return {}
    required = {"Sewing Line Code", "Target Morning", "Target Afternoon", "Target OT"}
    if not required.issubset(wip_df.columns):
        return {}

    targets = {}
    for _, row in wip_df.iterrows():
        code_raw = row.get("Sewing Line Code")
        if code_raw is None:
            continue
        code = str(code_raw).strip()
        if not code or not code.upper().startswith("VS"):
            continue
        m = pd.to_numeric(row.get("Target Morning"), errors="coerce")
        a = pd.to_numeric(row.get("Target Afternoon"), errors="coerce")
        ot = pd.to_numeric(row.get("Target OT"), errors="coerce")
        targets[code] = (
            float(m) if pd.notna(m) else 0.0,
            float(a) if pd.notna(a) else 0.0,
            float(ot) if pd.notna(ot) else 0.0,
        )

    return targets


# A JobCut - Suffix code is always exactly this shape: two digits, a dash,
# four digits, a dash, one digit (e.g. "26-0388-2") - always exactly the
# first 9 characters of a WIP Detail cell's line, whether or not a space
# follows it (e.g. "25-0621-1T.8,=60" is still "25-0621-1" + "T.8,=60").
_JOBCUT_PATTERN = re.compile(r"^\d{2}-\d{4}-\d")


def extract_jobcuts_from_detail(text) -> set:
    """Extract every JobCut - Suffix code mentioned in a WIP Detail Morning /
    Detail Afternoon / Detail OT cell's text. A cell can list multiple
    tables across multiple lines (e.g. "26-0388-2 T.8=180\\n26-0388-2
    T.11=180"); each line's own JobCut code is always its first 9
    characters. A line whose first 9 characters don't match that shape
    (blank, pure descriptive text, too short, etc.) is simply skipped
    rather than guessed at.

    Returns an empty set for blank/missing/malformed text - never raises.
    """
    if text is None:
        return set()
    text = str(text)
    if not text.strip():
        return set()

    found = set()
    for line in text.splitlines():
        line = line.strip()
        candidate = line[:9]
        if _JOBCUT_PATTERN.match(candidate):
            found.add(candidate)
    return found


def compute_wip_jobcut_restrictions(wip_df: pd.DataFrame) -> dict:
    """From an already-extracted Tab 2 WIP DataFrame (see load_wip()),
    build {Sewing Line code: set of JobCut - Suffix codes} - the JobCuts
    that line's own Detail Morning / Detail Afternoon / Detail OT cells
    actually mention. build_cut_plan() only plans, for each Sewing Line,
    tables whose JobCut - Suffix is in that line's own set here - a line
    with an EMPTY set (its Detail cells were blank, or mentioned nothing
    recognizable) has nothing planned for it, matching "only the JobCuts
    written in Detail Morning/Afternoon/OT should be planned".

    Matched the same way as compute_wip_session_targets() (via the "Sewing
    Line Code" column). Every Sewing Line Code found gets an entry in the
    returned dict - even one mapping to an empty set - so callers can tell
    "this line IS covered by Tab 2, it just has no JobCuts listed" apart
    from "Tab 2 doesn't cover this line at all" (a key that's simply
    absent), which is decided separately via compute_wip_session_targets().

    Returns {} if wip_df is None, empty, or missing the expected columns
    (e.g. an unstructured/raw-preview WIP file, or an older WIP template
    without Detail columns) - never raises.
    """
    if wip_df is None or len(wip_df) == 0:
        return {}
    detail_cols = ["Detail Morning", "Detail Afternoon", "Detail OT"]
    if "Sewing Line Code" not in wip_df.columns or not any(c in wip_df.columns for c in detail_cols):
        return {}

    restrictions = {}
    for _, row in wip_df.iterrows():
        code_raw = row.get("Sewing Line Code")
        code = str(code_raw).strip() if code_raw is not None else ""
        if not code or not code.upper().startswith("VS"):
            continue
        jobcuts = set()
        for col in detail_cols:
            if col in wip_df.columns:
                jobcuts |= extract_jobcuts_from_detail(row.get(col))
        restrictions.setdefault(code, set())
        restrictions[code] |= jobcuts

    return restrictions


# Sewing Lines that are known to have NO Morning/Afternoon/OT session-target
# breakdown in Tab 2 at all - only a single overall "Sewing Target Per Day"
# figure, which Tab 1's own extracted data already carries directly (see
# compute_fallback_day_target_plan() below). Sewing Lines not in this list
# are entirely unaffected: they're skipped from planning as usual (see rule
# 4 in build_cut_plan()) if Tab 2 doesn't cover them.
FALLBACK_TARGET_LINES = ["VSEW002", "VSEW006", "VSEW008", "VSEW011"]


def compute_fallback_day_target_plan(df: pd.DataFrame, lines=FALLBACK_TARGET_LINES):
    """For each Sewing Line in `lines` that appears in Tab 1's extracted
    data, build a single-session fallback plan using ONLY Tab 1's own
    "Sewing Target Per Day (with OT)" column - no Tab 2/WIP coverage
    required at all for these specific lines.

    Only the FIRST JobCut - Suffix to appear for that Sewing Line (in the
    source file's own row order - `df` is never re-sorted before this
    point, so `.iloc[0]` genuinely means "first in the file") gets planned;
    every other JobCut for that Sewing Line is excluded, matching "the
    model should extract/plan the JobCut that appears first in each Sewing
    Line". That JobCut's own Sewing Target Per Day (with OT) value (assumed
    constant across all its rows, as the source data has it) becomes the
    line's entire day target, all assigned to a single "Morning" session -
    Afternoon and OT both get a target of 0, since there's no session
    breakdown available for these lines and the request is simply "plan
    Qty in order to reach Sewing Target Per Day, defaulting everything to
    Morning". The existing cascading selection in build_cut_plan()/
    recalc_cut_plan() already handles a (target, 0, 0) triple correctly on
    its own - once cumulative Qty reaches that single target, it
    immediately exhausts the two zero-target sessions right behind it and
    the group stops - so no special-casing is needed there, only here.

    Returns (fallback_session_targets, fallback_jobcut_restrictions), each
    shaped exactly like compute_wip_session_targets()'s and
    compute_wip_jobcut_restrictions()'s own return values. Callers should
    only use these to FILL IN gaps left by real Tab 2 data (never override
    a Sewing Line Tab 2 already covers for real) - e.g. via dict.setdefault().
    A line missing from Tab 1's data entirely, or whose first JobCut has no
    usable numeric Sewing Target Per Day (with OT), simply gets no entry.
    """
    fallback_targets = {}
    fallback_restrictions = {}
    if df is None or len(df) == 0:
        return fallback_targets, fallback_restrictions

    sewing_line_str = df["Sewing Line"].astype(str)
    for line in lines:
        line_rows = df[sewing_line_str == line]
        if len(line_rows) == 0:
            continue
        first_jobcut = str(line_rows.iloc[0]["JobCut - Suffix"])
        jobcut_rows = line_rows[line_rows["JobCut - Suffix"].astype(str) == first_jobcut]
        day_target_numeric = pd.to_numeric(
            jobcut_rows.get("Sewing Target Per Day (with OT)"), errors="coerce"
        ).dropna()
        if len(day_target_numeric) == 0:
            continue
        fallback_targets[line] = (float(day_target_numeric.iloc[0]), 0.0, 0.0)
        fallback_restrictions[line] = {first_jobcut}

    return fallback_targets, fallback_restrictions


CUT_PLAN_COLUMNS = [
    "Sewing Line",
    "JobCut - Suffix",
    "Sewing Target Per Day",
    "Mark Type",
    "Table No.",
    "Sewing target Morning",
    "Cut Plan Morning",
    "Diff (Morning)",
    "Sewing target Afternoon",
    "Cut Plan Afternoon",
    "Diff (Afternoon)",
    "Sewing target OT",
    "Cut Plan OT",
    "Diff (OT)",
    "Note",
]

# The three (session name, target column, cut-plan column, diff column)
# tuples, in the fixed order tables get planned into: a group's tables are
# planned entirely into Morning until that session's running Diff turns
# non-negative, then Afternoon takes over the same way, then OT. See
# build_cut_plan()/recalc_cut_plan() for exactly how this drives selection.
CUT_PLAN_SESSIONS = [
    ("Morning", "Sewing target Morning", "Cut Plan Morning", "Diff (Morning)"),
    ("Afternoon", "Sewing target Afternoon", "Cut Plan Afternoon", "Diff (Afternoon)"),
    ("OT", "Sewing target OT", "Cut Plan OT", "Diff (OT)"),
]


def rows_to_cutplan_dataframe(rows) -> pd.DataFrame:
    """Convert a list of row dicts (as submitted from the editable Cut Plan
    table in the browser) back into a properly-typed plan DataFrame, ready
    to be written out with write_cut_plan_workbook()."""
    int_cols = {
        "Sewing Target Per Day",
        "Table No.",
        "Sewing target Morning",
        "Cut Plan Morning",
        "Diff (Morning)",
        "Sewing target Afternoon",
        "Cut Plan Afternoon",
        "Diff (Afternoon)",
        "Sewing target OT",
        "Cut Plan OT",
        "Diff (OT)",
    }

    def to_int(v):
        try:
            if v is None or v == "":
                return 0
            return int(float(v))
        except (TypeError, ValueError):
            return 0

    cleaned = []
    for row in rows:
        out = {}
        for col in CUT_PLAN_COLUMNS:
            val = row.get(col, "")
            out[col] = to_int(val) if col in int_cols else ("" if val is None else str(val))
        cleaned.append(out)

    return pd.DataFrame(cleaned, columns=CUT_PLAN_COLUMNS)


def _candidate_table_pool(df: pd.DataFrame, wip_jobcut_restrictions: Optional[dict] = None) -> pd.DataFrame:
    """Shared step for build_cut_plan() and recalc_cut_plan(): drop
    non-Pending colorway rows, drop rows whose JobCut isn't one Tab 2's
    Detail Morning/Afternoon/OT cells actually mention for that Sewing Line
    (when that restriction is available - see compute_wip_jobcut_restrictions()),
    then combine Qty across colorways sharing a Table ID. Returns one row
    per (Sewing Line, JobCut - Suffix, Mark Type, Table ID).

    Selection is purely status-based on top of that: only a colorway row
    whose Status is exactly "Pending" is eligible for planning -
    "Completed" AND "In Progress" are both excluded (not just "Completed").
    There is no separate date-based eligibility window on top of this (an
    earlier revision added a Start Cut / FG Start Date window; that's been
    removed since it was excluding tables that should still be planned).

    wip_jobcut_restrictions (optional): {Sewing Line code: set of allowed
    JobCut - Suffix codes}. When a Sewing Line HAS an entry here (even an
    empty set), only its listed JobCuts are eligible - a line whose Detail
    cells mentioned nothing that day has nothing planned for it. A Sewing
    Line with NO entry at all (key absent) is left unrestricted by this
    function - Tab 2 coverage for planning purposes is decided separately,
    by compute_wip_session_targets().

    Per-session Morning/Afternoon/OT targets are looked up separately per
    Sewing Line (see compute_wip_session_targets()) - this function no
    longer carries a single "Target" column, since planning now needs three
    numbers per group instead of one.
    """
    work = df.copy()
    work["Qty"] = pd.to_numeric(work["Qty"], errors="coerce").fillna(0)
    work["Table ID"] = pd.to_numeric(work["Table ID"], errors="coerce")

    status_norm = work["Status"].astype(str).str.strip().str.lower()
    work = work[status_norm == "pending"]

    if wip_jobcut_restrictions:
        def _jobcut_allowed(row):
            allowed = wip_jobcut_restrictions.get(row["Sewing Line"])
            return allowed is None or row["JobCut - Suffix"] in allowed

        work = work[work.apply(_jobcut_allowed, axis=1)]

    group_cols = ["Sewing Line", "JobCut - Suffix", "Mark Type"]
    table_cols = group_cols + ["Table ID"]

    table_level = (
        work.groupby(table_cols, dropna=False)
        .agg(
            Combined_Qty=("Qty", "sum"),
            Colorway_Count=("Colorway", "nunique"),
        )
        .reset_index()
    )
    return table_level


MERGE_COLUMNS = [
    "Sewing Line",
    "JobCut - Suffix",
    "Mark Type",
    "Sewing Target Per Day",
    "Sewing target Morning",
    "Sewing target Afternoon",
    "Sewing target OT",
]

# Which physical building each Sewing Line's cutting work happens in. Used to
# split the Cut Plan into two separate output tables/sheets: "Building 1"
# and "Building 2". A Sewing Line not listed here falls into "Unassigned"
# (kept, not dropped, so nothing silently disappears from the plan).
BUILDING_1_LINES = ["VSEW001", "VSEW002", "VSEW003", "VSEW004", "VSEW005", "VSEW007", "VSEW008", "VSEW015", "VSEW018", "VS02+06"]
BUILDING_2_LINES = ["VSEW006", "VSEW009", "VSEW010", "VSEW011", "VSEW012", "VSEW013", "VSEW014", "VSEW016", "VSEW017", "VS08+11"]


def get_building(sewing_line) -> str:
    """Map a Sewing Line to its building label."""
    s = str(sewing_line).strip()
    if s in BUILDING_1_LINES:
        return "Building 1"
    if s in BUILDING_2_LINES:
        return "Building 2"
    return "Unassigned"


def compute_multi_decoration_jobcuts(source_df: pd.DataFrame) -> set:
    """Find every (Sewing Line, JobCut - Suffix) that has MORE THAN ONE
    distinct Mark Type needing decoration (Decoration == "Yes" in the
    source data).

    These are the JobCuts where the planner does NOT have to plan every
    Mark Type, and does NOT have to plan enough tables to fully reach the
    Sewing Target Per Day - partial planning is a legitimate choice when a
    JobCut's decoration work is already split across multiple Mark Types.

    Returns a set of (sewing_line, jobcut_suffix) string tuples.
    """
    if "Decoration" not in source_df.columns:
        return set()

    needs_decoration = source_df[source_df["Decoration"].astype(str).str.strip().str.lower() == "yes"]
    if len(needs_decoration) == 0:
        return set()

    counts = needs_decoration.groupby(["Sewing Line", "JobCut - Suffix"])["Mark Type"].nunique()
    flagged = counts[counts > 1]
    return {(str(sl), str(jc)) for sl, jc in flagged.index}


def split_plan_by_building(plan_df: pd.DataFrame) -> "dict[str, pd.DataFrame]":
    """Split a Cut Plan DataFrame into per-building DataFrames, preserving
    row order within each. Only includes "Unassigned" if it's non-empty, so
    a fully-mapped plan never shows a stray empty section - but a Sewing
    Line that isn't in either building list is still kept, never dropped."""
    if len(plan_df) == 0:
        return {"Building 1": plan_df, "Building 2": plan_df}

    building_series = plan_df["Sewing Line"].map(get_building)
    result = {
        "Building 1": plan_df[building_series == "Building 1"].reset_index(drop=True),
        "Building 2": plan_df[building_series == "Building 2"].reset_index(drop=True),
    }
    unassigned = plan_df[building_series == "Unassigned"].reset_index(drop=True)
    if len(unassigned) > 0:
        result["Unassigned"] = unassigned
    return result


def _row_session(row):
    """Which of the three (Morning/Afternoon/OT) sessions this Cut Plan
    row belongs to, inferred from which pair of (Cut Plan X, Diff (X))
    columns is non-zero - each planned row belongs to exactly one session
    by construction (see build_cut_plan()/recalc_cut_plan()). Returns None
    for a row where none of the three pairs has anything yet (e.g. a
    freshly manually-added blank row the user hasn't filled in)."""
    def _nz(v):
        try:
            return float(v) != 0
        except (TypeError, ValueError):
            return False

    if _nz(row.get("Cut Plan Morning")) or _nz(row.get("Diff (Morning)")):
        return "Morning"
    if _nz(row.get("Cut Plan Afternoon")) or _nz(row.get("Diff (Afternoon)")):
        return "Afternoon"
    if _nz(row.get("Cut Plan OT")) or _nz(row.get("Diff (OT)")):
        return "OT"
    return None


def _session_target_merge_ranges(rows):
    """For each of the three session-target columns, compute the list of
    (start_idx, end_idx) row-index ranges (inclusive, 0-based into `rows`)
    that should be TRUE merged in the Excel output.

    Deliberately NOT derived from compute_continuation_flags()'s booleans -
    those only say "hide/show this one cell" and don't distinguish WHY a
    cell is hidden (a genuine continuation of the same group's same session,
    vs. simply being the wrong session for that column entirely). Naively
    merging every consecutive "hidden" run together would incorrectly span
    a merge across two unrelated blocks that just happen to both be hidden.
    This function instead walks the same (Mark Type group, session)
    structure directly, so a range is only ever produced for rows that are
    ACTUALLY the same group continuing the same session.
    """
    session_cols = {"Morning": "Sewing target Morning", "Afternoon": "Sewing target Afternoon", "OT": "Sewing target OT"}
    ranges = {col: [] for col in session_cols.values()}

    group_keys = [f"{r.get('Sewing Line', '')}|{r.get('JobCut - Suffix', '')}|{r.get('Mark Type', '')}" for r in rows]
    sessions = [_row_session(r) for r in rows]

    for session_name, col in session_cols.items():
        run_start = None
        run_group = None
        for idx in range(len(rows)):
            matches = sessions[idx] == session_name
            if matches and run_start is not None and group_keys[idx] == run_group:
                continue  # extend the current run
            # the current run (if any) has ended - close it out
            if run_start is not None and idx - 1 > run_start:
                ranges[col].append((run_start, idx - 1))
            run_start = idx if matches else None
            run_group = group_keys[idx] if matches else None
        if run_start is not None and len(rows) - 1 > run_start:
            ranges[col].append((run_start, len(rows) - 1))

    return ranges


def compute_continuation_flags(rows) -> list:
    """For each row, figure out - per merge-eligible column - whether it
    repeats the row above and should therefore be visually merged with it
    (blank/merged instead of shown again).

    Nesting rule:
      - Sewing Line and JobCut - Suffix are shown/hidden TOGETHER: both only
        merge with the row above when the JobCut - Suffix is unchanged. That
        way Sewing Line re-displays at the start of every JobCut, not just
        once for the whole Sewing Line block - easier to read which JobCut
        a given block belongs to.
      - Mark Type and Sewing Target Per Day merge with the row above only
        within the same JobCut AND (for Sewing Target Per Day) the same
        Mark Type - so they re-display at the start of every Mark Type
        sub-block.
      - Sewing target Morning/Afternoon/OT are shown ONLY on the first row
        of their OWN session within a Mark Type - e.g. Sewing target
        Afternoon shows on whichever row is the first to have Qty planned
        into Afternoon, not on the Mark Type's very first row. Every other
        row hides them, whether that other row is a LATER table in the
        same session (a true repeat) or a table in a DIFFERENT session
        entirely (which never had that target apply to it in the first
        place). build_cut_plan()/recalc_cut_plan() store the group's full
        target on every row regardless of session (consistent with how
        every other merge-eligible column works), so this function is what
        actually limits each target to appearing once, on the session's
        first table.
      - A session can be SKIPPED ENTIRELY - e.g. two tables together already
        cover all of Morning's and Afternoon's capacity, so the next table
        jumps straight into OT and no row ever "belongs to" Afternoon (see
        the multi-session skip-ahead in build_cut_plan()/recalc_cut_plan()).
        A skipped session's target must still be shown somewhere, even
        though its Qty planned is 0 for every row - it's shown alongside
        the target of whichever row is the first to move on to a LATER
        session, and/or on the group's very last row if the group ends
        before ever reaching a later session at all (this also covers a
        single-row group: Morning alone can cover the whole day, so
        Afternoon/OT are shown on that same row for context).

    rows: list of dicts (same shape as CUT_PLAN_COLUMNS rows), already
    sorted by Sewing Line -> JobCut - Suffix -> Mark Type -> Table No.

    Returns a list (same length as rows) of {column: bool} - True means
    "this cell repeats the row above, merge/blank it".
    """
    session_order = ["Morning", "Afternoon", "OT"]
    session_target_cols = {"Morning": "Sewing target Morning", "Afternoon": "Sewing target Afternoon", "OT": "Sewing target OT"}

    def group_key_of(row):
        return f"{row.get('Sewing Line', '')}|{row.get('JobCut - Suffix', '')}|{row.get('Mark Type', '')}"

    flags = [None] * len(rows)
    prev = None
    n = len(rows)
    i = 0
    while i < n:
        j = i
        while j < n and group_key_of(rows[j]) == group_key_of(rows[i]):
            j += 1
        # This Mark Type group spans rows[i:j]. last_shown_idx tracks how
        # far into session_order we've displayed a target so far (-1 = none
        # yet) so we can detect gaps (skipped sessions) as we go.
        last_shown_idx = -1
        for idx in range(i, j):
            row = rows[idx]
            row_flags = {}
            if prev is None:
                same_jobcut = False
                same_mark_type = False
            else:
                same_sewing_line = str(row.get("Sewing Line", "")) == str(prev.get("Sewing Line", ""))
                same_jobcut = same_sewing_line and str(row.get("JobCut - Suffix", "")) == str(prev.get("JobCut - Suffix", ""))
                same_mark_type = same_jobcut and str(row.get("Mark Type", "")) == str(prev.get("Mark Type", ""))

            row_flags["Sewing Line"] = same_jobcut
            row_flags["JobCut - Suffix"] = same_jobcut
            row_flags["Mark Type"] = same_mark_type
            row_flags["Sewing Target Per Day"] = same_mark_type

            for col in session_target_cols.values():
                row_flags[col] = True  # hidden by default; unhidden below as appropriate

            this_session = _row_session(row)
            this_idx = session_order.index(this_session) if this_session in session_order else None
            if this_idx is not None and this_idx > last_shown_idx:
                # First row to reach a session we haven't shown yet - show
                # its own target, plus any sessions skipped over to get here.
                for skip_idx in range(last_shown_idx + 1, this_idx + 1):
                    row_flags[session_target_cols[session_order[skip_idx]]] = False
                last_shown_idx = this_idx

            flags[idx] = row_flags
            prev = row

        # The group ended before ever reaching one or more of the later
        # sessions (e.g. everything fit in Morning alone) - show those
        # remaining targets on the group's last row so they aren't lost.
        if last_shown_idx < len(session_order) - 1:
            last_idx = j - 1
            for skip_idx in range(last_shown_idx + 1, len(session_order)):
                flags[last_idx][session_target_cols[session_order[skip_idx]]] = False

        i = j
    return flags


def build_cut_plan(
    df: pd.DataFrame,
    wip_session_targets: dict,
    run_datetime: Optional[datetime] = None,
    wip_jobcut_restrictions: Optional[dict] = None,
    manual_only_lines: Optional[set] = None,
):
    """Build the Tab 3 cut plan from Tab 1's extracted DataFrame, using
    Tab 2's per-Sewing-Line Morning/Afternoon/OT targets (see
    compute_wip_session_targets()) to decide how far into the day each
    group's tables get planned.

    Rules:
      1. Completed tables are excluded from planning. Status is recorded
         per colorway row, so a table with mixed colorway statuses only
         drops its already-completed colorway(s) — the remaining
         (Pending / In Progress) colorway(s) still get planned.
      1b. Only JobCuts mentioned in Tab 2's own Detail Morning / Detail
          Afternoon / Detail OT cells (see compute_wip_jobcut_restrictions())
          are planned for a given Sewing Line, when that restriction is
          available - a JobCut this line's Detail cells don't mention isn't
          planned, even if it has otherwise-eligible Pending tables. A line
          with an empty restriction set (its Detail cells were blank, or
          mentioned nothing recognizable) has nothing planned for it at
          all. A line with no restriction entry whatsoever (the WIP file
          doesn't include Detail columns, or wip_jobcut_restrictions isn't
          supplied) is left unrestricted by this rule.
      2. Multiple colorways can share the same Table ID within a
         (Sewing Line, JobCut - Suffix, Mark Type) group — their Qty is
         combined into one figure per table before planning.
      3. Within each group, tables are planned smallest Table ID first,
         one session at a time (Morning, then Afternoon, then OT):
           - Keep a running "Diff" = (total Qty planned in this group so
             far, across every table/session) minus (the sum of every
             session's target UP TO AND INCLUDING the one currently being
             filled). Diff starts using just the Morning target; the
             moment a table's addition pushes it non-negative (>= 0),
             that table is the LAST one for the current session — the
             NEXT table moves into the next eligible session, and the
             cumulative target used for comparison jumps up accordingly.
           - A single table's addition can clear MORE THAN ONE session's
             cumulative target at once (e.g. two mid-sized tables together
             already cover all of Morning's capacity AND all of
             Afternoon's) - when that happens, the next table skips
             straight past the now-fully-covered session(s) into whichever
             session the running total hasn't reached yet, rather than
             advancing just one session at a time regardless of how far
             the total has actually gone.
           - Once OT's cumulative target has also been reached, or all
             three sessions have been used, the group stops — any
             remaining tables for that group stay pending for a future
             run, not force-fit into an already-full day.
      3b. The group also stops as soon as its running total Planned Qty
          (summed across every table/session planned so far, including
          the table that just pushed it there) reaches or exceeds
          Sewing Target Per Day - even if a session boundary hasn't
          technically been crossed yet. This is a slightly tighter cap
          than rule 3's session-by-session cascading alone, since
          Sewing Target Per Day is the day's total rounded once (see the
          rounding note below), while the sum of the three individually
          rounded session targets used by rule 3 can occasionally be a
          couple of units higher.
      4. A Sewing Line with no matching row in Tab 2's WIP data is skipped
         entirely - no rows are planned for it at all this run, rather than
         falling back to some default. Tab 2 coverage is required for a
         line to be planned; run_info["lines_missing_wip_targets"] lists
         exactly which lines (present in Tab 1's data, candidates for
         planning) had no WIP coverage and were skipped, so the UI can
         surface this clearly instead of silently producing an
         incomplete-looking plan.

    Returns (plan_df, run_info).
    """
    table_level = _candidate_table_pool(df, wip_jobcut_restrictions=wip_jobcut_restrictions)
    fallback_targets_for_run_info, _ = compute_fallback_day_target_plan(df)

    group_cols = ["Sewing Line", "JobCut - Suffix", "Mark Type"]
    # Sort candidates by Sewing Line -> JobCut - Suffix -> Mark Type -> Table ID
    # up front, so the plan naturally comes out grouped/ordered that way too
    # (matches the reference planning sheet's layout).
    table_level = table_level.sort_values(group_cols + ["Table ID"]).reset_index(drop=True)
    output_rows = []
    lines_missing_targets = set()

    for keys, g in table_level.groupby(group_cols, sort=False, dropna=False):
        sewing_line, jobcut_suffix, mark_type = keys

        raw_targets = wip_session_targets.get(sewing_line)
        if raw_targets is None:
            # No Tab 2 coverage for this Sewing Line at all - skip planning
            # it entirely this run, rather than guessing with a (0, 0, 0)
            # fallback that would otherwise show every one of its tables as
            # immediately over target.
            lines_missing_targets.add(sewing_line)
            continue
        target_morning = math.ceil(raw_targets[0])
        target_afternoon = math.ceil(raw_targets[1])
        target_ot = math.ceil(raw_targets[2])
        # Round the TOTAL once, rather than summing the three already-rounded
        # session targets - ceil(a)+ceil(b)+ceil(c) can overshoot ceil(a+b+c)
        # by a couple of units purely from rounding each fractional part up
        # separately (e.g. Morning 144.0 + Afternoon 153.6 + OT 115.2 =
        # 412.8, which should round up to 413 - but 144+154+116 = 414).
        sewing_target_per_day = math.ceil(raw_targets[0] + raw_targets[1] + raw_targets[2])

        sessions = [
            (target_morning, "Cut Plan Morning", "Diff (Morning)"),
            (target_afternoon, "Cut Plan Afternoon", "Diff (Afternoon)"),
            (target_ot, "Cut Plan OT", "Diff (OT)"),
        ]
        cumulative_target_upto = []
        running_target = 0
        for t, _, _ in sessions:
            running_target += t
            cumulative_target_upto.append(running_target)

        def blank_row(table_id, qty_for_col=None):
            row = {
                "Sewing Line": sewing_line,
                "JobCut - Suffix": jobcut_suffix,
                "Sewing Target Per Day": int(sewing_target_per_day),
                "Mark Type": mark_type,
                "Table No.": table_id,
                "Sewing target Morning": int(target_morning),
                "Cut Plan Morning": 0,
                "Diff (Morning)": 0,
                "Sewing target Afternoon": int(target_afternoon),
                "Cut Plan Afternoon": 0,
                "Diff (Afternoon)": 0,
                "Sewing target OT": int(target_ot),
                "Cut Plan OT": 0,
                "Diff (OT)": 0,
                "Note": "",  # left blank - purely for the user's own manual comments
            }
            return row

        g_sorted = g.sort_values("Table ID", ascending=True).reset_index(drop=True)

        cumulative_qty = 0
        session_idx = 0

        for _, row in g_sorted.iterrows():
            if session_idx >= len(sessions):
                break  # all 3 sessions exhausted - remaining tables stay pending

            table_id = int(row["Table ID"]) if pd.notna(row["Table ID"]) else ""
            qty = int(row["Combined_Qty"])

            cumulative_qty += qty
            diff = cumulative_qty - cumulative_target_upto[session_idx]
            _, cut_col, diff_col = sessions[session_idx]

            out = blank_row(table_id)
            out[cut_col] = qty
            out[diff_col] = int(diff)
            output_rows.append(out)

            if diff >= 0:
                # This table's cumulative total didn't just clear the
                # CURRENT session's cumulative target - it may have
                # already run past one or more LATER sessions' cumulative
                # targets too (e.g. two medium tables together already
                # cover Morning AND Afternoon capacity). Keep advancing
                # past every session boundary the running total has
                # already cleared, so the NEXT table starts in the
                # correct (possibly further-ahead) session rather than
                # just the next one in line.
                while session_idx < len(sessions) and cumulative_qty >= cumulative_target_upto[session_idx]:
                    session_idx += 1

            if cumulative_qty >= sewing_target_per_day:
                # Planned Qty for this Mark Type has reached (or passed) the
                # day's total Sewing Target - stop the group right here,
                # even if a session boundary hasn't technically been
                # crossed yet (see rule 3b).
                break

    plan_df = pd.DataFrame(output_rows, columns=CUT_PLAN_COLUMNS)
    run_info = {
        "run_datetime": run_datetime or now_th(),
        "plan_date": (run_datetime or now_th()).date(),
        # Sewing Lines that had candidate tables in Tab 1's data but no
        # matching row in Tab 2's WIP data at all - these were skipped
        # entirely (see rule 4 above), so they'll never appear in plan_df
        # itself. Purely informational, so Tab 3's UI can flag exactly
        # which lines need Tab 2 coverage before they can be planned.
        "lines_missing_wip_targets": sorted(lines_missing_targets),
        # Carried through so recalc_cut_plan() can apply the SAME JobCut
        # restriction when re-selecting tables later (e.g. after the user
        # raises a target) - run_info is stored server-side only (never
        # round-tripped through the browser as JSON), so keeping this as a
        # plain dict of sets here is fine.
        "wip_jobcut_restrictions": wip_jobcut_restrictions or {},
        # The real (Morning, Afternoon, OT) numbers for any of
        # FALLBACK_TARGET_LINES present in this data (see
        # compute_fallback_day_target_plan()) - purely a defensive backstop
        # for the browser's swap-Diff calculation (see templates/
        # index.html's FALLBACK_SESSION_TARGETS) in the rare case a
        # manually-added row's Sewing target cells are genuinely blank
        # (e.g. a brand new group with no other row yet to supply a real
        # target) - every ordinarily-generated row already carries its own
        # real numeric target directly, so this is rarely needed.
        "fallback_session_targets": dict(fallback_targets_for_run_info),
        # Sewing Lines that GENUINELY needed the FALLBACK_TARGET_LINES
        # fallback (no real Tab 2 session-target coverage at all) - as
        # opposed to a FALLBACK_TARGET_LINES line that happens to ALSO
        # have real Tab 2 coverage, which should keep behaving completely
        # normally. Carried through so recalc_cut_plan() can treat every
        # row for these specific lines as permanently manual (see its
        # docstring) - editing their target/Qty only changes what's
        # displayed, it never automatically adds or drops tables, since
        # there's no genuine Tab 2 breakdown to justify doing so.
        "manual_only_lines": set(manual_only_lines or set()),
    }
    return plan_df, run_info


def lookup_table_qty(source_df: pd.DataFrame, sewing_line: str, jobcut_suffix: str, mark_type, table_id: int):
    """Look up a single table's TRUE combined quantity (colorways combined,
    non-Pending colorways excluded) straight from the source data - used when
    the user directly edits a row's Table No., so that row's quantity gets
    replaced with the real number for whichever table they typed in,
    instead of keeping the previous table's leftover value.

    Returns the quantity (int) if that exact (Sewing Line, JobCut - Suffix,
    Mark Type, Table ID) combination exists in the source data, else None.
    """
    table_level = _candidate_table_pool(source_df)
    match = table_level[
        (table_level["Sewing Line"].astype(str) == str(sewing_line))
        & (table_level["JobCut - Suffix"].astype(str) == str(jobcut_suffix))
        & (table_level["Mark Type"].astype(str) == str(mark_type))
        & (table_level["Table ID"] == table_id)
    ]
    if len(match) == 0:
        return None
    return int(match.iloc[0]["Combined_Qty"])


def recalc_cut_plan(source_df: pd.DataFrame, current_rows: list, run_info: dict):
    """Re-run table selection after the user has edited a Cut Plan Qty
    cell and/or a session target cell on the (already generated) Tab 3
    table.

    For each (Sewing Line, JobCut - Suffix, Mark Type) group present in
    current_rows:
      - Each of the three session targets (Morning/Afternoon/OT) is
        whatever the user's rows currently show for that column (max
        across the group's rows) - re-derived fresh each time, so editing
        any of them re-triggers the same cascading selection
        build_cut_plan() uses.
      - Each candidate table's quantity is the user's edited quantity for
        that table (summed across whichever of the three Cut Plan
        columns it's showing something in) if that table is currently on
        the page, otherwise the original combined Qty from the source
        data (tables the user never saw/touched).
      - Tables are re-selected smallest Table ID first, using the exact
        same session-by-session cascading rule build_cut_plan() uses -
        this can both ADD tables (a target raised, or an edited qty
        lowered the running total) and DROP tables (a target lowered, or
        an edited qty raised the running total past it sooner). The
        group also still stops as soon as its running total Planned Qty
        reaches Sewing Target Per Day, same as build_cut_plan()'s rule 3b.
      - Manual comments already typed into Note are preserved for tables
        that remain selected.
      - A row the browser flags as manually added ("+ Add row" or the
        per-row insert button in templates/index.html) is ALWAYS kept
        exactly as the user currently has it, whether its Table No.
        matches a real, otherwise-eligible Pending table or is entirely
        made up - "+ Add row" is meant to let the user plan an extra
        table by hand on top of whatever the automatic cascade already
        decided, not to resubmit that table for the cascade to accept or
        reject based on whether the group has "room" left. A manually
        flagged real table is excluded from the automatic candidate pool
        so it isn't ALSO picked up a second time by the normal selection.
      - Rows for tables the user manually added (a Table No. that doesn't
        exist in the source data for that group) are always kept exactly
        as typed - their own Cut Plan Morning/Afternoon/OT and Diff values
        are never recomputed.
      - The SAME JobCut restriction build_cut_plan() applied (see
        compute_wip_jobcut_restrictions()) is re-applied here too, carried
        forward via run_info["wip_jobcut_restrictions"] - a JobCut a
        Sewing Line's Detail cells didn't mention still can't be pulled in
        by raising a target, for consistency with the original plan.
      - Sewing Lines with no genuine Tab 2 session breakdown (see
        FALLBACK_TARGET_LINES / compute_fallback_day_target_plan()) have
        EVERY one of their rows treated as manually-kept, regardless of
        that row's own "_manual" flag - these lines' Morning/Afternoon/OT
        targets are only ever a single Sewing Target Per Day figure being
        used as a stand-in, not a real per-session breakdown from Tab 2,
        so there's no genuine basis for automatically adding or dropping
        tables when the user edits it. Editing the target (or a Cut Plan
        Qty) for one of these lines only ever changes what's displayed on
        the rows already there - the initial set of tables from
        build_cut_plan() stays exactly as generated until the user adds or
        removes tables by hand.

    Returns a fresh plan_df (same columns/shape as build_cut_plan's output).
    """
    wip_jobcut_restrictions = run_info.get("wip_jobcut_restrictions") or {}
    table_level = _candidate_table_pool(source_df, wip_jobcut_restrictions=wip_jobcut_restrictions)
    if "manual_only_lines" in run_info:
        # Prefer what build_cut_plan() already determined (it can tell a
        # FALLBACK_TARGET_LINES line that ALSO has real Tab 2 coverage
        # apart from one that genuinely needed the fallback - this
        # function only has source_df, which can't make that distinction).
        manual_only_lines = run_info["manual_only_lines"]
    else:
        manual_only_lines, _ = compute_fallback_day_target_plan(source_df)
        manual_only_lines = set(manual_only_lines.keys())

    # Index candidates for fast lookup: (Sewing Line, JobCut - Suffix, Mark Type) -> DataFrame.
    # Keys are normalized to strings on BOTH sides (here, and again below when
    # building rows_by_group from the browser's posted rows) - the source
    # data's own dtypes (e.g. Mark Type usually comes through as a numeric
    # int64 from the Excel column) otherwise silently mismatch every string
    # the browser ever sends (HTML <input> values are always strings), which
    # would make EVERY row look like it has no matching candidate group at
    # all - falling through to being treated as an unmatched manual row
    # regardless of whether the user actually added it by hand.
    group_cols = ["Sewing Line", "JobCut - Suffix", "Mark Type"]
    candidates_by_group = {
        tuple(str(k) for k in keys): g.reset_index(drop=True)
        for keys, g in table_level.groupby(group_cols, sort=False, dropna=False)
    }

    def to_num(v, default=0.0):
        try:
            if v is None or v == "":
                return default
            return float(v)
        except (TypeError, ValueError):
            return default

    # Organize the user's current rows by group. Keys are normalized to
    # strings (see the matching comment on candidates_by_group above) so a
    # numeric-vs-string type mismatch with the source data's dtypes can
    # never cause a group to look like it has no real candidates.
    rows_by_group = {}
    for row in current_rows:
        key = (str(row.get("Sewing Line", "")), str(row.get("JobCut - Suffix", "")), str(row.get("Mark Type", "")))
        rows_by_group.setdefault(key, []).append(row)

    output_rows = []

    for key, user_rows in rows_by_group.items():
        sewing_line, jobcut_suffix, mark_type = key
        # Round UP for the same reason as build_cut_plan() - the user could
        # type in a fractional target directly, and it should never make
        # the group look like it needs less than it actually does.
        target_morning = math.ceil(max((to_num(r.get("Sewing target Morning")) for r in user_rows), default=0.0))
        target_afternoon = math.ceil(max((to_num(r.get("Sewing target Afternoon")) for r in user_rows), default=0.0))
        target_ot = math.ceil(max((to_num(r.get("Sewing target OT")) for r in user_rows), default=0.0))
        # Preserve whatever's already in the user's current rows for Sewing
        # Target Per Day, rather than re-deriving it from the three
        # (already-rounded) session targets above - this function only ever
        # sees whole-number session targets (no raw fractional WIP data to
        # work from), so summing them would reproduce the same over-rounding
        # bug build_cut_plan() avoids by rounding the raw total once. The
        # value was already computed correctly there; recalculating a plan
        # shouldn't silently drift it to a different number.
        sewing_target_per_day = int(max((to_num(r.get("Sewing Target Per Day")) for r in user_rows), default=0.0))

        candidates = candidates_by_group.get(key)
        if sewing_line in manual_only_lines:
            # This line has no genuine Tab 2 session-target breakdown to
            # justify automatically adding tables (see this function's
            # docstring) - empty the candidate pool entirely so the normal
            # cascading loop below has nothing left to auto-select from,
            # not just excluding the tables already on the page. Without
            # this, any OTHER real Pending table for this group that
            # simply wasn't part of the original selection could still get
            # silently pulled in by the cascade the moment enough "room"
            # opened up (e.g. after lowering the target) - the whole point
            # here is that nothing gets added or dropped automatically at
            # all for these lines.
            candidates = None

        # Build override map: Table ID (int) -> user's current Qty for that
        # row (summed across whichever of the three Cut Plan columns it's
        # showing something in), and preserve any Note text already on
        # that row. Also collect rows whose Table No. isn't a real
        # candidate for this group — manually added extras that don't
        # exist in the source data at all. A row only counts as an
        # override if it has an EXPLICIT quantity somewhere; if all three
        # Cut Plan columns are blank (e.g. the user just changed which
        # Table No. this row points to, so its old values no longer mean
        # anything), it's treated as having no known quantity yet, and
        # falls back to that table's real quantity from the source data
        # instead of incorrectly reusing a stale value.
        #
        # A row the browser flags "_manual" (created via "+ Add row" or
        # the per-row insert button - see buildRow()/collectAllRows() in
        # templates/index.html) is ALWAYS kept exactly as the user has it,
        # even when its Table No. matches a perfectly real, otherwise-
        # eligible Pending table - "+ Add row" is explicitly meant to let
        # the user plan an EXTRA table by hand, on top of whatever the
        # automatic cascade already decided, not to re-offer that same
        # table up for the normal cascade to accept-or-reject based on
        # whether the group has "room" left. Its Table ID is excluded from
        # the normal candidate pool below so it isn't also picked up a
        # second time by the automatic selection.
        qty_override = {}
        note_by_table = {}
        manual_rows = []
        manual_table_ids = set()
        for r in user_rows:
            table_no_raw = r.get("Table No.", "")
            try:
                table_id = int(float(table_no_raw))
            except (TypeError, ValueError):
                table_id = None

            is_real_candidate = (
                candidates is not None and table_id is not None
                and (candidates["Table ID"] == table_id).any()
            )
            is_manual = bool(r.get("_manual")) or sewing_line in manual_only_lines

            if is_manual:
                manual_rows.append(r)
                if is_real_candidate:
                    manual_table_ids.add(table_id)
                continue

            morning_raw = r.get("Cut Plan Morning", "")
            afternoon_raw = r.get("Cut Plan Afternoon", "")
            ot_raw = r.get("Cut Plan OT", "")
            has_known_qty = not (
                (morning_raw is None or str(morning_raw).strip() == "")
                and (afternoon_raw is None or str(afternoon_raw).strip() == "")
                and (ot_raw is None or str(ot_raw).strip() == "")
            )

            if is_real_candidate:
                note_by_table[table_id] = r.get("Note", "")
                if has_known_qty:
                    qty_override[table_id] = to_num(morning_raw) + to_num(afternoon_raw) + to_num(ot_raw)
            else:
                manual_rows.append(r)

        if manual_table_ids and candidates is not None:
            candidates = candidates[~candidates["Table ID"].isin(manual_table_ids)]

        sessions = [
            (target_morning, "Cut Plan Morning", "Diff (Morning)"),
            (target_afternoon, "Cut Plan Afternoon", "Diff (Afternoon)"),
            (target_ot, "Cut Plan OT", "Diff (OT)"),
        ]
        cumulative_target_upto = []
        running_target = 0
        for t, _, _ in sessions:
            running_target += t
            cumulative_target_upto.append(running_target)

        selected_rows_out = []

        if candidates is not None and len(candidates) > 0:
            g_sorted = candidates.sort_values("Table ID", ascending=True).reset_index(drop=True)
            cumulative_qty = 0
            session_idx = 0

            for _, crow in g_sorted.iterrows():
                if session_idx >= len(sessions):
                    break  # all 3 sessions exhausted - remaining tables stay pending

                table_id = int(crow["Table ID"]) if pd.notna(crow["Table ID"]) else None
                qty = int(qty_override.get(table_id, crow["Combined_Qty"]))

                blank = {
                    "table_id": table_id,
                    "Cut Plan Morning": 0, "Diff (Morning)": 0,
                    "Cut Plan Afternoon": 0, "Diff (Afternoon)": 0,
                    "Cut Plan OT": 0, "Diff (OT)": 0,
                }

                cumulative_qty += qty
                diff = cumulative_qty - cumulative_target_upto[session_idx]
                _, cut_col, diff_col = sessions[session_idx]
                blank[cut_col] = qty
                blank[diff_col] = int(diff)
                selected_rows_out.append(blank)

                if diff >= 0:
                    # See build_cut_plan()'s matching comment: keep
                    # advancing past every session boundary the running
                    # total has already cleared, not just the current one.
                    while session_idx < len(sessions) and cumulative_qty >= cumulative_target_upto[session_idx]:
                        session_idx += 1

                if cumulative_qty >= sewing_target_per_day:
                    # Planned Qty for this Mark Type has reached (or passed)
                    # the day's total Sewing Target - stop the group right
                    # here, mirroring build_cut_plan()'s rule 3b.
                    break

        for row_out in selected_rows_out:
            table_id = row_out["table_id"]
            note = note_by_table.get(table_id, "")
            output_rows.append(
                {
                    "Sewing Line": sewing_line,
                    "JobCut - Suffix": jobcut_suffix,
                    "Sewing Target Per Day": int(sewing_target_per_day),
                    "Mark Type": mark_type,
                    "Table No.": table_id,
                    "Sewing target Morning": int(target_morning),
                    "Cut Plan Morning": row_out["Cut Plan Morning"],
                    "Diff (Morning)": row_out["Diff (Morning)"],
                    "Sewing target Afternoon": int(target_afternoon),
                    "Cut Plan Afternoon": row_out["Cut Plan Afternoon"],
                    "Diff (Afternoon)": row_out["Diff (Afternoon)"],
                    "Sewing target OT": int(target_ot),
                    "Cut Plan OT": row_out["Cut Plan OT"],
                    "Diff (OT)": row_out["Diff (OT)"],
                    "Note": note,
                    "_manual": False,
                }
            )

        for r in manual_rows:
            output_rows.append(
                {
                    "Sewing Line": sewing_line,
                    "JobCut - Suffix": jobcut_suffix,
                    "Sewing Target Per Day": int(sewing_target_per_day),
                    "Mark Type": mark_type,
                    "Table No.": r.get("Table No.", ""),
                    "Sewing target Morning": int(target_morning),
                    "Cut Plan Morning": int(to_num(r.get("Cut Plan Morning"))),
                    "Diff (Morning)": int(to_num(r.get("Diff (Morning)"))),
                    "Sewing target Afternoon": int(target_afternoon),
                    "Cut Plan Afternoon": int(to_num(r.get("Cut Plan Afternoon"))),
                    "Diff (Afternoon)": int(to_num(r.get("Diff (Afternoon)"))),
                    "Sewing target OT": int(target_ot),
                    "Cut Plan OT": int(to_num(r.get("Cut Plan OT"))),
                    "Diff (OT)": int(to_num(r.get("Diff (OT)"))),
                    "Note": r.get("Note", ""),
                    # Kept True for every row that ever passed through here
                    # (both real-table overrides and genuinely fictitious
                    # manually-typed Table Nos.), so the browser keeps
                    # treating it as a manual, always-preserved row across
                    # any FUTURE recalculation too (see buildRow()/
                    # collectAllRows() in templates/index.html).
                    "_manual": True,
                }
            )

    # "_manual" rides along as an extra, non-schema column purely for the
    # browser's benefit (so it can keep marking a row as user-added across
    # repeated recalculations) - CUT_PLAN_COLUMNS itself (the Excel/Save
    # schema) is untouched.
    result_df = pd.DataFrame(output_rows, columns=CUT_PLAN_COLUMNS + ["_manual"])
    if len(result_df) > 0:
        # Sort by Sewing Line -> JobCut - Suffix -> Mark Type, stable so each
        # group's already-correct internal Table No. order is preserved.
        result_df = result_df.sort_values(group_cols, kind="stable").reset_index(drop=True)
    return result_df


# Documents the Tab 3 planning rules in plain English. No longer written
# into the downloaded Excel files (the "Logic & Assumptions" sheet was
# removed) - kept here purely as an in-code reference for anyone reading
# the source, describing the same logic implemented in build_cut_plan()/
# recalc_cut_plan() below.
CUT_PLAN_ASSUMPTIONS_TEXT = [
    "CUT PLANNING MODEL v2 — TAB 3: CUT PLAN",
    "",
    "1. INPUT",
    "Built from Tab 1's extracted Buffer Cutting Order Form data (already filtered to Sewing",
    "Lines starting with \"VS\"), using Tab 2's (WIP Upload) per-Sewing-Line Morning/Afternoon/OT",
    "targets to decide how far into the day each group's tables get planned. A Sewing Line with",
    "no matching row in Tab 2's WIP data is skipped entirely this run - Tab 2 coverage is",
    "required for a line to be planned.",
    "",
    "2. COMPLETED TABLES ARE EXCLUDED",
    "Status is recorded per colorway row. Any row with Status = \"Completed\" is dropped before",
    "planning. A table whose colorways have mixed status (e.g. one colorway Completed, another",
    "still Pending) is NOT excluded entirely — only its completed colorway's quantity is dropped;",
    "the remaining colorway(s) are still planned.",
    "",
    "3. COMBINING COLORWAYS",
    "A single Table ID can have more than one colorway. Their Qty is summed into one combined",
    "figure per table before planning.",
    "",
    "3b. NOTE COLUMN",
    "Left blank by the model. It's only filled in if the user types something into it manually",
    "on the editable Cut Plan tab (e.g. a manual adjustment reason) before saving.",
    "",
    "4. TABLE SELECTION RULE: THREE SESSIONS, CASCADING",
    "Within each (Sewing Line, JobCut - Suffix, Mark Type) group, tables are planned SMALLEST",
    "Table ID first, one session at a time - Morning, then Afternoon, then OT. A running Diff =",
    "(total Qty planned in the group so far, across every table/session) minus (the sum of every",
    "session's target up to and including the one currently being filled). The moment a table's",
    "addition pushes that Diff to zero or above, that table is the LAST one for the current",
    "session - the next table moves into the next eligible session. If that table's addition was",
    "large enough to clear MORE THAN ONE session's cumulative target at once (e.g. two mid-sized",
    "tables together already cover all of Morning's capacity AND all of Afternoon's), the next",
    "table skips straight past the now-fully-covered session(s) into whichever session the running",
    "total hasn't reached yet - it doesn't advance just one session at a time regardless of how far",
    "the total has actually gone. Once OT's cumulative target has also been reached, or all three",
    "sessions have been used, the group stops - any remaining tables for that group stay pending",
    "for a future run rather than being force-fit into an already-full day.",
    "",
    "4b. STOP ONCE PLANNED QTY REACHES THE DAY'S TOTAL TARGET",
    "The group also stops as soon as its running total Planned Qty (summed across every table/",
    "session planned so far, including the table that just pushed it there) reaches or exceeds",
    "Sewing Target Per Day - even if a session boundary under rule 4 hasn't technically been",
    "crossed yet. This applies the same way to every Job Cut and Mark Type.",
    "",
    "5. CUT PLAN MORNING/AFTERNOON/OT AND DIFF (MORNING)/(AFTERNOON)/(OT)",
    "Each table belongs to exactly ONE session - its own Qty appears in that session's Cut Plan",
    "column, with the other two sessions' Cut Plan columns left at 0 for that row. The",
    "corresponding Diff column shows the running Diff at that point (see rule 4).",
]


def _write_cut_plan_sheet(ws, plan_df: pd.DataFrame, title_text: str, edited_at: Optional[datetime] = None) -> None:
    """Write one Cut Plan table (one building's worth of rows) into the given
    worksheet, starting directly with the header row - no title/notice rows
    above it. title_text/edited_at are still accepted (existing callers
    pass them) but no longer written into the sheet itself; the run
    date/time and "manually edited" state are shown in the web UI instead."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
    cell_font = Font(name=FONT_NAME, size=10)
    note_font = Font(name=FONT_NAME, size=10, italic=True, color=WARN_FONT_COLOR)
    group_fill_a = PatternFill("solid", fgColor="DCE6F1")
    group_fill_b = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="B7B7B7")
    thick_top = Side(style="thick", color="1F4E78")
    medium_top = Side(style="medium", color="7C99B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_new_sewing_line = Border(left=thin, right=thin, top=thick_top, bottom=thin)
    border_new_jobcut = Border(left=thin, right=thin, top=medium_top, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    header_row = 1
    headers = list(plan_df.columns)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    if len(plan_df) > 0:
        rows_as_dicts = plan_df.to_dict(orient="records")
        continuation_flags = compute_continuation_flags(rows_as_dicts)

        group_key_series = (
            plan_df["Sewing Line"].astype(str)
            + "|"
            + plan_df["JobCut - Suffix"].astype(str)
            + "|"
            + plan_df["Mark Type"].astype(str)
        )
        sewing_line_series = plan_df["Sewing Line"].astype(str)
        jobcut_series = plan_df["Sewing Line"].astype(str) + "|" + plan_df["JobCut - Suffix"].astype(str)
        current_key = None
        fill_toggle = False
        prev_sewing_line = None
        prev_jobcut = None

        col_index = {h: j for j, h in enumerate(headers, start=1)}

        for i, row in enumerate(plan_df.itertuples(index=False), start=header_row + 1):
            idx = i - (header_row + 1)
            key = group_key_series.iloc[idx]
            if key != current_key:
                fill_toggle = not fill_toggle
                current_key = key
            fill = group_fill_a if fill_toggle else group_fill_b

            sewing_line_val = sewing_line_series.iloc[idx]
            jobcut_val = jobcut_series.iloc[idx]
            if sewing_line_val != prev_sewing_line and prev_sewing_line is not None:
                row_border = border_new_sewing_line
            elif jobcut_val != prev_jobcut and prev_jobcut is not None:
                row_border = border_new_jobcut
            else:
                row_border = border
            prev_sewing_line = sewing_line_val
            prev_jobcut = jobcut_val

            row_flags = continuation_flags[idx]
            values = list(row)
            for j, val in enumerate(values, start=1):
                col_name = headers[j - 1]
                is_continuation = col_name in MERGE_COLUMNS and row_flags.get(col_name, False)
                c = ws.cell(row=i, column=j, value=(None if is_continuation else val))
                c.border = row_border
                c.fill = fill
                if col_name == "Note":
                    c.font = note_font
                    c.alignment = left
                else:
                    c.font = cell_font
                    c.alignment = center

        # ---- Apply true cell merges for repeated values in MERGE_COLUMNS ----
        # The three session-target columns need their OWN merge-range logic
        # (see _session_target_merge_ranges()) rather than the simple
        # "any consecutive hidden run merges with the row above it" approach
        # used for the other merge columns below. That simple approach
        # breaks for these three specifically: a row hidden because it's the
        # WRONG session for that column can sit right next to a row hidden
        # because it's a genuine continuation of a DIFFERENT group's same
        # session - naively merging every consecutive "hidden" run together
        # would incorrectly span the merge across a Mark Type/group boundary
        # it has nothing to do with.
        session_target_cols = {"Sewing target Morning", "Sewing target Afternoon", "Sewing target OT"}
        simple_merge_cols = [c for c in MERGE_COLUMNS if c not in session_target_cols]

        for col_name in simple_merge_cols:
            j = col_index[col_name]
            run_start = None
            for idx, flags in enumerate(continuation_flags):
                excel_row = header_row + 1 + idx
                if flags.get(col_name, False):
                    continue
                if run_start is not None:
                    run_end = header_row + idx
                    if run_end > run_start:
                        ws.merge_cells(start_row=run_start, start_column=j, end_row=run_end, end_column=j)
                run_start = excel_row
            if run_start is not None:
                run_end = header_row + len(continuation_flags)
                if run_end > run_start:
                    ws.merge_cells(start_row=run_start, start_column=j, end_row=run_end, end_column=j)

        for col_name, ranges in _session_target_merge_ranges(rows_as_dicts).items():
            j = col_index[col_name]
            for start_idx, end_idx in ranges:
                ws.merge_cells(
                    start_row=header_row + 1 + start_idx, start_column=j,
                    end_row=header_row + 1 + end_idx, end_column=j,
                )

    widths = [12, 16, 16, 10, 8, 14, 12, 11, 14, 12, 11, 14, 12, 11, 46]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = f"A{header_row + 1}"
    last_row = max(len(plan_df) + header_row, header_row)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"


def write_single_building_workbook(
    building_df: pd.DataFrame,
    run_info: dict,
    building_label: str,
    downloaded_date,
    output_path: str,
    edited_at: Optional[datetime] = None,
) -> None:
    """Write ONE standalone workbook for a single building: just its Cut
    Plan sheet (titled "<building_label> of <downloaded_date>").

    downloaded_date should be the date the file is actually being generated
    for download (i.e. "today"), not necessarily the plan's target date -
    pass a date object.
    """
    wb = Workbook()
    wb.remove(wb.active)

    generated_at = run_info["run_datetime"].strftime("%Y-%m-%d %H:%M")
    ws = wb.create_sheet("Cut Plan")
    title_text = f"{building_label} of {downloaded_date.isoformat()}   |   Generated: {generated_at}"
    _write_cut_plan_sheet(ws, building_df, title_text, edited_at=edited_at)

    wb.save(output_path)


def write_cut_plan_workbook(
    plan_df: pd.DataFrame,
    run_info: dict,
    output_path: str,
    edited_at: Optional[datetime] = None,
) -> None:
    """Write the formatted Tab 3 output workbook: one sheet per building
    ("Building 1", "Building 2", and "Unassigned" if any Sewing Line doesn't
    map to either).

    Pass edited_at when this call is regenerating the file after the user
    manually adjusted the plan table in the browser - it's noted in each
    sheet's title row so it's clear the numbers no longer come straight from
    the automatic selection rule.
    """
    wb = Workbook()
    wb.remove(wb.active)  # replaced by the per-building sheets below

    generated_at = run_info["run_datetime"].strftime("%Y-%m-%d %H:%M")
    plan_date_str = run_info["plan_date"].isoformat()

    by_building = split_plan_by_building(plan_df)
    # Keep a stable, sensible sheet order even when a building is empty.
    ordered_keys = [k for k in ["Building 1", "Building 2", "Unassigned"] if k in by_building]

    for key in ordered_keys:
        building_df = by_building[key]
        ws = wb.create_sheet(key)
        title_text = f"{key} of {plan_date_str}   |   Generated: {generated_at}"
        _write_cut_plan_sheet(ws, building_df, title_text, edited_at=edited_at)

    wb.save(output_path)
